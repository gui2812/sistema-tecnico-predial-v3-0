from __future__ import annotations

import base64
import json
import re
import shutil
import subprocess
import tempfile
import unicodedata
from copy import copy
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pypdf import PdfReader, PdfWriter


BASE_DIR = Path(__file__).resolve().parent.parent

TEMPLATE_EXCEL = (
    BASE_DIR
    / "public"
    / "templates"
    / "Modelo - Mapa de Cotação.xlsx"
)

NOME_PLANILHA = "Mapa de Cotação"
MAX_ITENS = 9

app = FastAPI(
    title="Sistema Técnico Predial — Backend",
    version="3.3.3",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================================================
# UTILIDADES
# =========================================================

def texto(
    valor: Any,
) -> str:
    return str(
        valor
        or ""
    ).strip()


def nome_seguro(
    valor: Any,
) -> str:
    valor = texto(
        valor
    )

    valor = unicodedata.normalize(
        "NFD",
        valor,
    )

    valor = "".join(
        caractere
        for caractere in valor
        if unicodedata.category(
            caractere
        ) != "Mn"
    )

    valor = re.sub(
        r'[\\/:*?"<>|]',
        "-",
        valor,
    )

    valor = re.sub(
        r"\s+",
        " ",
        valor,
    )

    return valor.strip()


def numero_decimal(
    valor: Any,
) -> float:
    if isinstance(
        valor,
        (
            int,
            float,
        ),
    ):
        return float(
            valor
        )

    valor = texto(
        valor
    )

    if not valor:
        return 0.0

    valor = re.sub(
        r"[^\d,.\-]",
        "",
        valor,
    )

    if "," in valor:
        valor = valor.replace(
            ".",
            "",
        )

        valor = valor.replace(
            ",",
            ".",
        )

    try:
        return float(
            valor
        )

    except ValueError:
        return 0.0


def valor_frete_excel(
    valor: Any,
) -> Any:
    valor_texto = texto(
        valor
    )

    if not valor_texto:
        return "N/A"

    possui_numero = bool(
        re.search(
            r"\d",
            valor_texto,
        )
    )

    if not possui_numero:
        return valor_texto

    return numero_decimal(
        valor_texto
    )


def valor_frete_numerico(
    valor: Any,
) -> float:
    valor_texto = texto(
        valor
    )

    if not valor_texto:
        return 0.0

    possui_numero = bool(
        re.search(
            r"\d",
            valor_texto,
        )
    )

    if not possui_numero:
        return 0.0

    return numero_decimal(
        valor_texto
    )


def data_br(
    valor: Any,
) -> str:
    valor = texto(
        valor
    )

    if not valor:
        return ""

    partes = valor.split(
        "-"
    )

    if len(
        partes
    ) == 3:
        return (
            f"{partes[2]}"
            f"/{partes[1]}"
            f"/{partes[0]}"
        )

    return valor


def codigo_mapa(
    dados: dict[str, Any],
) -> str:
    ano = texto(
        dados.get(
            "ano"
        )
    )

    numero = texto(
        dados.get(
            "numeroMapa"
        )
    )

    numero = re.sub(
        r"\s+",
        "",
        numero,
    )

    if (
        ano
        and numero.startswith(
            ano
        )
    ):
        return numero

    return (
        f"{ano}"
        f"{numero}"
    )


def nome_base_mapa(
    dados: dict[str, Any],
) -> str:
    fornecedores = (
        dados.get(
            "fornecedores"
        )
        or []
    )

    empresa = (
        nome_seguro(
            dados.get(
                "empresaAprovada"
            )
        )
        or nome_seguro(
            fornecedores[0].get(
                "empresa"
            )
            if fornecedores
            else ""
        )
        or "Fornecedor"
    )

    identificacao = (
        nome_seguro(
            dados.get(
                "identificacaoMapa"
            )
        )
        or "Cotacao"
    )

    return (
        f"Mapa {codigo_mapa(dados)}"
        f" - {empresa}"
        f" - {identificacao}"
    )


def serializar_arquivo_base64(
    caminho: Path,
) -> dict[str, str]:
    conteudo = caminho.read_bytes()

    return {
        "nome":
            caminho.name,

        "base64":
            base64.b64encode(
                conteudo
            ).decode(
                "ascii"
            ),
    }


# =========================================================
# ITENS
# =========================================================

def normalizar_itens(
    dados: dict[str, Any],
) -> list[dict[str, Any]]:
    itens_recebidos = (
        dados.get(
            "itens"
        )
        or []
    )

    if (
        isinstance(
            itens_recebidos,
            list,
        )
        and itens_recebidos
    ):
        return [
            {
                "descricao":
                    texto(
                        item.get(
                            "descricao"
                        )
                    ),

                "quantidade":
                    numero_decimal(
                        item.get(
                            "quantidade"
                        )
                    ),

                "unidade":
                    texto(
                        item.get(
                            "unidade"
                        )
                    ),
            }
            for item in itens_recebidos[
                :MAX_ITENS
            ]
        ]

    # Compatibilidade com mapas antigos.
    return [
        {
            "descricao":
                texto(
                    dados.get(
                        "descricaoItem"
                    )
                ),

            "quantidade":
                numero_decimal(
                    dados.get(
                        "quantidade"
                    )
                ),

            "unidade":
                texto(
                    dados.get(
                        "unidade"
                    )
                ),
        }
    ]


def obter_preco_item(
    fornecedor: dict[str, Any],
    indice_item: int,
) -> dict[str, Any]:
    precos_itens = (
        fornecedor.get(
            "precosItens"
        )
        or []
    )

    if (
        indice_item
        < len(
            precos_itens
        )
    ):
        return (
            precos_itens[
                indice_item
            ]
            or {}
        )

    # Compatibilidade com versão antiga.
    if (
        indice_item
        == 0
    ):
        return {
            "precoUnitario":
                fornecedor.get(
                    "precoUnitario"
                ),

            "precoTotal":
                fornecedor.get(
                    "precoTotal"
                ),
        }

    return {}


def obter_preco_unitario_item(
    fornecedor: dict[str, Any],
    item: dict[str, Any],
    indice_item: int,
) -> float:
    preco = obter_preco_item(
        fornecedor,
        indice_item,
    )

    preco_unitario = numero_decimal(
        preco.get(
            "precoUnitario"
        )
    )

    preco_total = numero_decimal(
        preco.get(
            "precoTotal"
        )
    )

    quantidade = numero_decimal(
        item.get(
            "quantidade"
        )
    )

    if (
        preco_unitario == 0
        and preco_total > 0
        and quantidade > 0
    ):
        return (
            preco_total
            / quantidade
        )

    return preco_unitario


def calcular_subtotal_item(
    fornecedor: dict[str, Any],
    item: dict[str, Any],
    indice_item: int,
) -> float:
    quantidade = numero_decimal(
        item.get(
            "quantidade"
        )
    )

    preco_unitario = obter_preco_unitario_item(
        fornecedor,
        item,
        indice_item,
    )

    return (
        quantidade
        * preco_unitario
    )


def calcular_total_fornecedor(
    fornecedor: dict[str, Any],
    itens: list[dict[str, Any]],
) -> float:
    subtotais = sum(
        calcular_subtotal_item(
            fornecedor,
            item,
            indice_item,
        )
        for indice_item, item in enumerate(
            itens
        )
    )

    return (
        subtotais
        + valor_frete_numerico(
            fornecedor.get(
                "frete"
            )
        )
    )


# =========================================================
# FORMATAÇÃO
# =========================================================

def alinhar_sem_perder_estilo(
    ws,
    endereco: str,
    horizontal: str,
    quebrar_texto: bool = False,
) -> None:
    alinhamento = copy(
        ws[
            endereco
        ].alignment
    )

    alinhamento.horizontal = (
        horizontal
    )

    alinhamento.vertical = (
        "center"
    )

    alinhamento.wrap_text = (
        quebrar_texto
    )

    ws[
        endereco
    ].alignment = alinhamento


def aplicar_formatacao_dinamica(
    ws,
) -> None:
    # Conta Orçamentária:
    # título e valor à esquerda.

    alinhar_sem_perder_estilo(
        ws,
        "B8",
        horizontal="left",
        quebrar_texto=True,
    )

    alinhar_sem_perder_estilo(
        ws,
        "E8",
        horizontal="left",
        quebrar_texto=True,
    )

    # Observações:
    # título no canto esquerdo.

    alinhar_sem_perder_estilo(
        ws,
        "B30",
        horizontal="left",
        quebrar_texto=True,
    )

    # Texto da observação:
    # centralizado na área branca.

    alinhar_sem_perder_estilo(
        ws,
        "B31",
        horizontal="center",
        quebrar_texto=True,
    )

    # Empresa aprovada:
    # centralizada.

    alinhar_sem_perder_estilo(
        ws,
        "O37",
        horizontal="center",
        quebrar_texto=True,
    )


# =========================================================
# BLOCOS DE APROVAÇÃO
# =========================================================

def garantir_blocos_aprovacoes(
    ws,
) -> None:
    """
    Recria os três quadros de aprovação usando células e bordas.

    Os quadros passam a fazer parte da estrutura da planilha,
    permanecendo visíveis no Excel e no PDF gerado.
    """

    borda_cinza = Side(
        style="thin",
        color="A6A6A6",
    )

    sem_borda = Side(
        style=None,
    )

    preenchimento_branco = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF",
    )

    fonte_legenda = Font(
        name="Arial",
        size=9,
        color="666666",
    )

    quadros = [
        {
            "coluna_inicial": 2,   # B
            "coluna_final": 4,     # D
            "coluna_legenda": 3,   # C
            "legenda": "Contratante",
        },
        {
            "coluna_inicial": 6,   # F
            "coluna_final": 8,     # H
            "coluna_legenda": 7,   # G
            "legenda": "Gerência",
        },
        {
            "coluna_inicial": 10,  # J
            "coluna_final": 12,    # L
            "coluna_legenda": 11,  # K
            "legenda": "Diretoria",
        },
    ]

    linha_inicial = 37
    linha_final = 40

    for quadro in quadros:
        coluna_inicial = quadro[
            "coluna_inicial"
        ]

        coluna_final = quadro[
            "coluna_final"
        ]

        for linha in range(
            linha_inicial,
            linha_final + 1,
        ):
            for coluna in range(
                coluna_inicial,
                coluna_final + 1,
            ):
                celula = ws.cell(
                    row=linha,
                    column=coluna,
                )

                celula.fill = (
                    preenchimento_branco
                )

                celula.border = Border(
                    left=(
                        borda_cinza
                        if coluna
                        == coluna_inicial
                        else sem_borda
                    ),

                    right=(
                        borda_cinza
                        if coluna
                        == coluna_final
                        else sem_borda
                    ),

                    top=(
                        borda_cinza
                        if linha
                        == linha_inicial
                        else sem_borda
                    ),

                    bottom=(
                        borda_cinza
                        if linha
                        == linha_final
                        else sem_borda
                    ),
                )

        celula_legenda = ws.cell(
            row=linha_final,
            column=quadro[
                "coluna_legenda"
            ],
        )

        celula_legenda.value = (
            quadro[
                "legenda"
            ]
        )

        celula_legenda.font = (
            fonte_legenda
        )

        celula_legenda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


# =========================================================
# FORNECEDORES
# =========================================================

def limpar_fornecedor(
    ws,
    indice: int,
) -> None:
    colunas = [
        "L",
        "O",
        "R",
    ]

    coluna = colunas[
        indice
    ]

    for linha in [
        5,
        6,
        7,
        8,
        9,
    ]:
        ws[
            f"{coluna}{linha}"
        ] = ""

    for linha in range(
        12,
        21,
    ):
        ws[
            f"{coluna}{linha}"
        ] = 0

    ws[
        f"{coluna}22"
    ] = "N/A"

    ws[
        f"{coluna}23"
    ] = ""

    ws[
        f"{coluna}24"
    ] = ""

    ws[
        f"{coluna}25"
    ] = ""

    ws[
        f"{coluna}26"
    ] = ""


def preencher_fornecedor(
    ws,
    fornecedor: dict[str, Any],
    indice: int,
    itens: list[dict[str, Any]],
) -> None:
    colunas = [
        "L",
        "O",
        "R",
    ]

    coluna = colunas[
        indice
    ]

    ws[
        f"{coluna}5"
    ] = texto(
        fornecedor.get(
            "empresa"
        )
    )

    ws[
        f"{coluna}6"
    ] = texto(
        fornecedor.get(
            "contato"
        )
    )

    ws[
        f"{coluna}7"
    ] = texto(
        fornecedor.get(
            "telefone"
        )
    )

    ws[
        f"{coluna}8"
    ] = texto(
        fornecedor.get(
            "email"
        )
    )

    ws[
        f"{coluna}9"
    ] = data_br(
        fornecedor.get(
            "dataProposta"
        )
    )

    # Preços unitários dos itens:
    # linhas 12 até 20.

    for indice_item, item in enumerate(
        itens
    ):
        linha = (
            12
            + indice_item
        )

        ws[
            f"{coluna}{linha}"
        ] = obter_preco_unitario_item(
            fornecedor,
            item,
            indice_item,
        )

    # Limpa linhas que não foram utilizadas.

    for linha in range(
        12
        + len(
            itens
        ),
        21,
    ):
        ws[
            f"{coluna}{linha}"
        ] = 0

    ws[
        f"{coluna}22"
    ] = valor_frete_excel(
        fornecedor.get(
            "frete"
        )
    )

    ws[
        f"{coluna}23"
    ] = texto(
        fornecedor.get(
            "prazoEntrega"
        )
    )

    ws[
        f"{coluna}24"
    ] = texto(
        fornecedor.get(
            "validadeProposta"
        )
    )

    ws[
        f"{coluna}25"
    ] = texto(
        fornecedor.get(
            "condicaoPagamento"
        )
    )

    ws[
        f"{coluna}26"
    ] = texto(
        fornecedor.get(
            "garantia"
        )
    )


# =========================================================
# EXCEL PRINCIPAL PARA DOWNLOAD
# =========================================================

def preencher_excel(
    dados: dict[str, Any],
    caminho_saida: Path,
) -> None:
    if not TEMPLATE_EXCEL.exists():
        raise RuntimeError(
            "Template Excel não encontrado."
        )

    workbook = load_workbook(
        TEMPLATE_EXCEL
    )

    ws = workbook[
        NOME_PLANILHA
    ]

    # Preserva os dados fixos existentes no template:
    #
    # E5 = Empreendimento
    # E6 = Departamento
    # E7 = Contratante
    # E9 = Gerência Responsável

    # Conta Orçamentária.

    ws[
        "E8"
    ] = texto(
        dados.get(
            "contaOrcamentaria"
        )
    )

    itens = normalizar_itens(
        dados
    )

    # Linhas disponíveis no template:
    # linha 12 até linha 20.

    for indice_item in range(
        MAX_ITENS
    ):
        linha = (
            12
            + indice_item
        )

        if (
            indice_item
            < len(
                itens
            )
        ):
            item = itens[
                indice_item
            ]

            ws[
                f"B{linha}"
            ] = item[
                "descricao"
            ]

            ws[
                f"I{linha}"
            ] = item[
                "quantidade"
            ]

            ws[
                f"J{linha}"
            ] = item[
                "unidade"
            ]

        else:
            ws[
                f"B{linha}"
            ] = ""

            ws[
                f"I{linha}"
            ] = ""

            ws[
                f"J{linha}"
            ] = ""

    fornecedores = (
        dados.get(
            "fornecedores"
        )
        or []
    )[
        :3
    ]

    for indice in range(
        3
    ):
        if (
            indice
            < len(
                fornecedores
            )
        ):
            preencher_fornecedor(
                ws,
                fornecedores[
                    indice
                ],
                indice,
                itens,
            )

        else:
            limpar_fornecedor(
                ws,
                indice,
            )

    ws[
        "B31"
    ] = texto(
        dados.get(
            "observacoes"
        )
    )

    ws[
        "O37"
    ] = texto(
        dados.get(
            "empresaAprovada"
        )
    )

    aplicar_formatacao_dinamica(
        ws
    )

    garantir_blocos_aprovacoes(
        ws
    )

    workbook.calculation.fullCalcOnLoad = (
        True
    )

    workbook.calculation.forceFullCalc = (
        True
    )

    workbook.calculation.calcMode = (
        "auto"
    )

    workbook.save(
        caminho_saida
    )


# =========================================================
# EXCEL TEMPORÁRIO USADO SOMENTE PARA GERAR O PDF
# =========================================================

def preparar_excel_para_pdf(
    caminho_excel_origem: Path,
    caminho_excel_pdf: Path,
    dados: dict[str, Any],
) -> None:
    """
    O PDF nasce exclusivamente de uma cópia do Excel preenchido.

    A cópia temporária recebe subtotais e totais consolidados
    numericamente para evitar divergências de recálculo durante
    a conversão automática pelo LibreOffice.
    """

    shutil.copy2(
        caminho_excel_origem,
        caminho_excel_pdf,
    )

    workbook = load_workbook(
        caminho_excel_pdf
    )

    ws = workbook[
        NOME_PLANILHA
    ]

    itens = normalizar_itens(
        dados
    )

    fornecedores = (
        dados.get(
            "fornecedores"
        )
        or []
    )[
        :3
    ]

    colunas_preco_unitario = [
        "L",
        "O",
        "R",
    ]

    colunas_subtotal = [
        "M",
        "P",
        "S",
    ]

    celulas_total = [
        "L28",
        "O28",
        "R28",
    ]

    for indice_fornecedor in range(
        3
    ):
        fornecedor = (
            fornecedores[
                indice_fornecedor
            ]
            if indice_fornecedor
            < len(
                fornecedores
            )
            else {}
        )

        coluna_unitario = (
            colunas_preco_unitario[
                indice_fornecedor
            ]
        )

        coluna_subtotal = (
            colunas_subtotal[
                indice_fornecedor
            ]
        )

        for indice_item in range(
            MAX_ITENS
        ):
            linha = (
                12
                + indice_item
            )

            if (
                indice_item
                < len(
                    itens
                )
            ):
                item = itens[
                    indice_item
                ]

                unitario = obter_preco_unitario_item(
                    fornecedor,
                    item,
                    indice_item,
                )

                subtotal = calcular_subtotal_item(
                    fornecedor,
                    item,
                    indice_item,
                )

            else:
                unitario = 0
                subtotal = 0

            ws[
                f"{coluna_unitario}{linha}"
            ] = unitario

            ws[
                f"{coluna_subtotal}{linha}"
            ] = subtotal

        ws[
            celulas_total[
                indice_fornecedor
            ]
        ] = calcular_total_fornecedor(
            fornecedor,
            itens,
        )

    aplicar_formatacao_dinamica(
        ws
    )

    garantir_blocos_aprovacoes(
        ws
    )

    workbook.save(
        caminho_excel_pdf
    )


# =========================================================
# CONVERSÃO DO EXCEL PARA PDF
# =========================================================

def converter_excel_para_pdf(
    caminho_excel: Path,
    pasta_saida: Path,
) -> Path:
    pasta_perfil = (
        pasta_saida
        / "libreoffice-profile"
    )

    pasta_perfil.mkdir(
        parents=True,
        exist_ok=True,
    )

    comando = [
        "libreoffice",
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        (
            "-env:UserInstallation="
            f"file://{pasta_perfil}"
        ),
        "--convert-to",
        "pdf:calc_pdf_Export",
        "--outdir",
        str(
            pasta_saida
        ),
        str(
            caminho_excel
        ),
    ]

    resultado = subprocess.run(
        comando,
        check=False,
        capture_output=True,
        text=True,
        timeout=120,
    )

    caminho_pdf = (
        pasta_saida
        / f"{caminho_excel.stem}.pdf"
    )

    if (
        resultado.returncode != 0
        or not caminho_pdf.exists()
    ):
        raise RuntimeError(
            "Falha ao converter o Excel preenchido para PDF. "
            f"Saída: {resultado.stdout} "
            f"Erro: {resultado.stderr}"
        )

    return caminho_pdf


# =========================================================
# UNIÃO DOS PDFS
# =========================================================

def unir_pdfs(
    pdf_mapa: Path,
    pdfs_propostas: list[Path],
    caminho_saida: Path,
) -> None:
    writer = PdfWriter()

    caminhos = [
        pdf_mapa,
        *pdfs_propostas,
    ]

    for caminho in caminhos:
        try:
            reader = PdfReader(
                str(
                    caminho
                )
            )

            for pagina in reader.pages:
                writer.add_page(
                    pagina
                )

        except Exception as erro:
            raise RuntimeError(
                (
                    "Não foi possível ler o PDF: "
                    f"{caminho.name}"
                )
            ) from erro

    with caminho_saida.open(
        "wb"
    ) as arquivo_saida:
        writer.write(
            arquivo_saida
        )


async def salvar_upload_pdf(
    upload: UploadFile | None,
    pasta: Path,
    indice: int,
) -> Path | None:
    if (
        upload is None
        or not upload.filename
    ):
        return None

    if not upload.filename.lower().endswith(
        ".pdf"
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                f"A proposta {indice} "
                "precisa ser enviada em PDF."
            ),
        )

    destino = (
        pasta
        / f"proposta-{indice}.pdf"
    )

    with destino.open(
        "wb"
    ) as arquivo_destino:
        shutil.copyfileobj(
            upload.file,
            arquivo_destino,
        )

    return destino


# =========================================================
# ROTAS
# =========================================================

@app.get("/")
def inicio():
    return {
        "servico":
            "Sistema Técnico Predial — Backend",

        "status":
            "online",

        "versao":
            "3.3.3",

        "itens_por_mapa":
            MAX_ITENS,

        "geracao_pdf":
            "excel-preenchido-convertido",

        "aprovacoes":
            "quadros-recriados-com-celulas-e-bordas",
    }


@app.get(
    "/api/health"
)
def health_check():
    return {
        "status":
            "ok",

        "template_excel_encontrado":
            TEMPLATE_EXCEL.exists(),

        "template_excel":
            str(
                TEMPLATE_EXCEL
            ),

        "geracao_pdf":
            "somente-a-partir-do-excel",

        "itens_por_mapa":
            MAX_ITENS,

        "observacao":
            "titulo-esquerda-conteudo-centralizado",

        "conta_orcamentaria":
            "titulo-e-valor-a-esquerda",

        "aprovacoes":
            "quadros-recriados-com-celulas-e-bordas",
    }


@app.post(
    "/api/mapa-cotacao/gerar"
)
async def gerar_mapa_cotacao(
    dados_json: str = Form(...),
    proposta1: UploadFile | None = File(None),
    proposta2: UploadFile | None = File(None),
    proposta3: UploadFile | None = File(None),
):
    try:
        dados = json.loads(
            dados_json
        )

    except json.JSONDecodeError as erro:
        raise HTTPException(
            status_code=400,
            detail="Dados do formulário inválidos.",
        ) from erro

    with tempfile.TemporaryDirectory() as pasta_temporaria:
        pasta = Path(
            pasta_temporaria
        )

        nome_base = nome_base_mapa(
            dados
        )

        identificacao = (
            nome_seguro(
                dados.get(
                    "identificacaoMapa"
                )
            )
            or "Cotacao"
        )

        caminho_excel = (
            pasta
            / f"{nome_base}.xlsx"
        )

        preencher_excel(
            dados,
            caminho_excel,
        )

        caminho_excel_pdf = (
            pasta
            / (
                f"{nome_base}"
                " - conversao-pdf.xlsx"
            )
        )

        preparar_excel_para_pdf(
            caminho_excel,
            caminho_excel_pdf,
            dados,
        )

        caminho_pdf_convertido = converter_excel_para_pdf(
            caminho_excel_pdf,
            pasta,
        )

        caminho_pdf_mapa = (
            pasta
            / f"{nome_base}.pdf"
        )

        shutil.copy2(
            caminho_pdf_convertido,
            caminho_pdf_mapa,
        )

        propostas_upload = [
            proposta1,
            proposta2,
            proposta3,
        ]

        pdfs_propostas = []

        for indice, upload in enumerate(
            propostas_upload,
            start=1,
        ):
            pdf = await salvar_upload_pdf(
                upload,
                pasta,
                indice,
            )

            if pdf:
                pdfs_propostas.append(
                    pdf
                )

        caminho_pdf_completo = (
            pasta
            / (
                f"Mapa {codigo_mapa(dados)}"
                f" - {identificacao}"
                " + Prop.pdf"
            )
        )

        unir_pdfs(
            caminho_pdf_mapa,
            pdfs_propostas,
            caminho_pdf_completo,
        )

        return {
            "nomeBase":
                nome_base,

            "excel":
                serializar_arquivo_base64(
                    caminho_excel
                ),

            "pdfMapa":
                serializar_arquivo_base64(
                    caminho_pdf_mapa
                ),

            "pdfCompleto":
                serializar_arquivo_base64(
                    caminho_pdf_completo
                ),
        }
